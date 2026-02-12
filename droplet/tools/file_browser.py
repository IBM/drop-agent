"""File browser tool for exploring the local filesystem"""

import asyncio
import html
import os
from urllib.parse import unquote, urlparse

from gpt_oss.tools.simple_browser.backend import BackendError
from gpt_oss.tools.simple_browser.page_contents import process_html
from gpt_oss.tools.simple_browser.simple_browser_tool import SimpleBrowserTool
from openai_harmony import ToolNamespaceConfig
from openpyxl import load_workbook
from pypdf import PdfReader


class FileBackend:
    """
    Backend for browsing the local file system.

    - Directories return listing of contents
    - Files return content (supports text, PDF, Excel)
    """

    source = "files"

    def __init__(self, max_chars=32000):
        self.max_chars = max_chars

    async def search(self, query, topn=10, session=None):
        """
        List directory contents.
        topn parameter is ignored for file browsing (kept for interface compatibility).
        """
        if not os.path.isdir(query):
            raise BackendError(f"Path is not a directory: {query}")
        return await self._list_directory(query)

    async def fetch(self, path, session=None):
        """
        Fetch local file or directory content.
        """
        # Extract path from file:// URL or use as-is for absolute paths
        if path.startswith("file://"):
            if path.startswith("file://./"):
                path = unquote(path[7:])
            else:
                parsed = urlparse(path)
                if parsed.netloc in ('.', '..'):
                    path = unquote(parsed.netloc + parsed.path)
                else:
                    path = unquote(parsed.path)

        abs_path = os.path.abspath(path)

        if not os.path.exists(abs_path):
            raise BackendError(f"Path does not exist: {abs_path}")

        if os.path.isdir(abs_path):
            return await self._list_directory(abs_path)
        else:
            return await self._read_file(abs_path)

    async def _list_directory(self, path):
        """
        List directory contents and return as HTML.
        """
        def _sync_walk():
            files = []
            hidden_folders = []
            venv_folders = []

            for root, dirs, filenames in os.walk(path):
                hidden_dirs = [d for d in dirs if d.startswith(".")]
                hidden_folders.extend([os.path.join(root, d) for d in hidden_dirs])

                venv_dirs = []
                for d in dirs:
                    dir_path = os.path.join(root, d)
                    if os.path.isfile(os.path.join(dir_path, "pyvenv.cfg")) or \
                       os.path.isdir(os.path.join(dir_path, "conda-meta")):
                        venv_dirs.append(d)
                venv_folders.extend([os.path.join(root, d) for d in venv_dirs])

                dirs[:] = [d for d in dirs if not d.startswith(".") and d not in venv_dirs]

                for filename in filenames:
                    files.append(os.path.join(root, filename))

            return {
                "files": files,
                "hidden_folders": hidden_folders,
                "venv_folders": venv_folders
            }

        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, _sync_walk)

        # Format as HTML
        files_html = "\n".join([f"<li><a href=\"file://{f}\">{f}</a></li>" for f in result["files"]])
        hidden_html = "\n".join([f"<li>{f} (hidden)</li>" for f in result["hidden_folders"]])
        venv_html = "\n".join([f"<li>{f} (venv)</li>" for f in result["venv_folders"]])

        html_page = f"""
<html>
<head><title>Directory: {html.escape(path)}</title></head>
<body>
<h1>Directory Contents: {html.escape(path)}</h1>
<h2>Files ({len(result["files"])})</h2>
<ul>
{files_html}
</ul>
<h2>Hidden Folders ({len(result["hidden_folders"])})</h2>
<ul>
{hidden_html}
</ul>
<h2>Virtual Environments ({len(result["venv_folders"])})</h2>
<ul>
{venv_html}
</ul>
</body>
</html>
"""

        return process_html(
            html=html_page,
            url=f"file://{path}",
            title=f"Directory: {path}",
            display_urls=False,
            session=None,
        )

    async def _read_file(self, file_path):
        """
        Read file contents and return as HTML.
        """
        def _sync_read():
            abs_path = os.path.abspath(file_path)

            # PDF files
            if abs_path.lower().endswith('.pdf'):
                reader = PdfReader(abs_path)
                content = ""
                for page in reader.pages:
                    content += page.extract_text() + "\n"

                original_size = len(content)
                cropped = False
                warning = None

                if len(content) > self.max_chars:
                    content = content[:self.max_chars]
                    cropped = True
                    warning = f"WARNING: PDF content was cropped. Original size: {original_size} chars, showing first {self.max_chars} chars"

                return {
                    "file_path": abs_path,
                    "content": content,
                    "file_type": "pdf",
                    "cropped": cropped,
                    "warning": warning
                }

            # Excel files
            if abs_path.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                workbook = load_workbook(abs_path, read_only=True, data_only=True)
                content = ""

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    content += f"## Sheet: {sheet_name}\n\n"

                    rows = list(sheet.iter_rows(values_only=True))
                    if not rows:
                        content += "*Empty sheet*\n\n"
                        continue

                    header = rows[0]
                    col_count = len(header)

                    header_str = "| " + " | ".join(str(cell) if cell is not None else "" for cell in header) + " |"
                    content += header_str + "\n"

                    separator = "|" + "|".join(["-" * 3 for _ in range(col_count)]) + "|"
                    content += separator + "\n"

                    for row in rows[1:]:
                        row_str = "| " + " | ".join(str(cell) if cell is not None else "" for cell in row) + " |"
                        content += row_str + "\n"

                    content += "\n"

                workbook.close()

                original_size = len(content)
                cropped = False
                warning = None

                if len(content) > self.max_chars:
                    content = content[:self.max_chars]
                    cropped = True
                    warning = f"WARNING: Excel content was cropped. Original size: {original_size} chars, showing first {self.max_chars} chars"

                return {
                    "file_path": abs_path,
                    "content": content,
                    "file_type": "excel",
                    "cropped": cropped,
                    "warning": warning
                }

            # Check for binary files
            with open(abs_path, 'rb') as f:
                chunk = f.read(8192)
                if b'\x00' in chunk:
                    return {
                        "file_path": abs_path,
                        "error": "Binary file format is not supported",
                        "content": None,
                        "file_type": "binary"
                    }

            # Text files
            with open(abs_path, 'r', encoding='utf-8') as f:
                content = f.read()

            original_size = len(content)
            cropped = False
            warning = None

            if len(content) > self.max_chars:
                content = content[:self.max_chars]
                cropped = True
                warning = f"WARNING: File content was cropped. Original size: {original_size} chars, showing first {self.max_chars} chars"

            return {
                "file_path": abs_path,
                "content": content,
                "file_type": "text",
                "cropped": cropped,
                "warning": warning
            }

        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, _sync_read)

        if result.get("error"):
            raise BackendError(result["error"])

        # Format as HTML
        warning_html = f"<p><strong>{html.escape(result['warning'])}</strong></p>" if result.get("warning") else ""
        content_escaped = html.escape(result["content"])

        html_page = f"""
<html>
<head><title>File: {html.escape(result["file_path"])}</title></head>
<body>
<h1>File: {html.escape(result["file_path"])}</h1>
<p>Type: {result["file_type"]}</p>
{warning_html}
<pre>{content_escaped}</pre>
</body>
</html>
"""

        return process_html(
            html=html_page,
            url=f"file://{result['file_path']}",
            title=f"File: {result['file_path']}",
            display_urls=False,
            session=None,
        )


class FileBrowserTool(SimpleBrowserTool):
    """
    Browser tool for exploring the local file system.

    Uses the "files" namespace with standard SimpleBrowserTool functions.
    Just wraps a FileBackend that formats filesystem data as HTML for browsing.

    Provides:
    - search: List directory contents
    - open: Read file contents (text, PDF, Excel)
    - find: Find text patterns in the current page
    """

    def __init__(self):
        # SimpleBrowserTool asserts name=="browser", so we pass that
        # and override the name property below
        super().__init__(backend=FileBackend(), name="browser")

    @classmethod
    def get_tool_name(cls) -> str:
        return "files"

    @property
    def name(self) -> str:
        return self.get_tool_name()

    @property
    def tool_config(self):
        # Use native gpt-oss browser config structure but customize descriptions for file browsing
        config = ToolNamespaceConfig.browser()
        config.name = "files"
        config.description = """Tool for browsing the local file system.
The `cursor` appears in brackets before each browsing display: `[{cursor}]`.
Use search(query=".") to list the current directory.
Use search(query="/path/to/dir") to list a specific directory.
Use open(id=N) to open a file from search results, where N is the link number.
Cite information from the tool using the following format:
`【{cursor}†L{line_start}(-L{line_end})?】`, for example: `【6†L9-L11】` or `【8†L3】`.
Do not quote more than 10 words directly from the tool output.
sources=""" + self.backend.source

        # Customize the search function description for file browsing
        # Convert to dict to modify
        config_dict = config.model_dump()
        for tool_def in config_dict['tools']:
            if tool_def['name'] == 'search':
                tool_def['description'] = 'List directory contents. Use query="." for current directory or query="/path" for a specific directory.'
                tool_def['parameters']['properties']['query']['description'] = 'Directory path to list (use "." for current directory)'

        # Recreate config from modified dict
        config = ToolNamespaceConfig(**config_dict)

        return config
